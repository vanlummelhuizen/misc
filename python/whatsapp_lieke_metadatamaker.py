#!python

"""
TODO
"""

from jinja2 import Environment, select_autoescape, FileSystemLoader
from openpyxl import load_workbook
import sys
import os
from lxml import etree
import re

if len(sys.argv) != 5:
    print("""Usage:
    %s <metadata-excel> <template-file> <output-dir> <folia-file-dir>
    """ % sys.argv[0])
    exit()


FIRST_ROW_IN_METADATA_XSLX = 1
NUMBER_OF_ROWS_TO_PROCESS = 4
METADATA_COLUMNS = [
    'file_name_orig',
    'file_name_anonymous',
    'age',
    'sex'
]
OUTPUT_FILE_EXTENSION = 'cmdi.xml'
FOLIA_FILE_EXTENSION = 'folia.xml'

metadata_excel_file = sys.argv[1]
template_file_path = sys.argv[2]
output_dir = sys.argv[3]
folia_dir = sys.argv[4]


def count_words_in_folia(folia_file):
    try:
        xmlparser = etree.XMLParser(encoding='ISO-8859-1')
        root = etree.parse(folia_file, xmlparser).getroot()
        count = root.xpath("count(//f:w[@class=\"WORD\" or @class=\"PUNCTUATION\" or @class=\"Vern\"])",
                           namespaces={'f':'http://ilk.uvt.nl/folia'})
        return int(count)
    except:
        sys.stderr.write("Folia file %s could not analysed.\n" % folia_file)



# --- Prepare Jinja2 ---

template_file_name = os.path.basename(template_file_path)
template_dir_name = os.path.dirname(template_file_path)

env = Environment(
    autoescape=select_autoescape(['html', 'xml']),
    loader=FileSystemLoader(template_dir_name),
    trim_blocks=False
)
template = env.get_template(template_file_name)


# --- Read Metadata Excel ---

# Load the workbook
workbook = load_workbook(filename=metadata_excel_file)
# Take the first worksheet
worksheet = workbook[workbook.get_sheet_names()[0]]

# max_row = NUMBER_OF_ROWS_TO_PROCESS + FIRST_ROW_IN_METADATA_XSLX
max_row = worksheet.max_row+1

for row_index in range(FIRST_ROW_IN_METADATA_XSLX, max_row):
    row_index_str = str(row_index)

    # Put the data of a row in a convenient dictionary
    row_data = [cell.value for cell in worksheet['A'+row_index_str:'D'+row_index_str][0]]
    row_dict = dict(list(zip(METADATA_COLUMNS, row_data)))

    file_name = row_dict['file_name_anonymous']
    file_name = re.sub(r'^Onbewerkt\/', '', file_name)
    file_name = re.sub(r'\.$', '', file_name)

    sex = ""
    if row_dict['sex'] == 'meisje':
        sex = "female"
    elif row_dict['sex'] == 'jongen':
        sex = "male"

    word_count = count_words_in_folia(os.path.sep.join((folia_dir, file_name + "." + FOLIA_FILE_EXTENSION)))

    output_file = os.path.sep.join((output_dir, file_name + '.' + OUTPUT_FILE_EXTENSION))

    with open(output_file, 'w') as file:
        file.write(template.render(
            project_name='ACAD',
            text_type='WR-U-E-A_chats',
            published='N',
            publisher='',
            # publication_date_yyyy_mm_dd=row_dict['date'].strftime("%Y-%m-%d"),
            # genre=row_dict['genre'],
            # author_name=row_dict['authors'],
            author_sex=sex,
            author_age=row_dict['age'],
            resource_proxy_id=file_name,
            number_of_word_tokens=word_count or ''
        ))
        file.close()
