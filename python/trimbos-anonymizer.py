#!/usr/bin/python

"""
TODO
"""

import sys
import os
import getopt
import re
from filecollectionprocessing.fileprocessor import FileProcessor
from filecollectionprocessing.filecollectionprocessor import FileCollectionProcessor
from openpyxl import load_workbook

class TrimbosAnonymizer(FileProcessor):
    _extensions = ["xml", "xls", "xlsx"]
    _agent_ids_csv = 'agent_id.csv'

    def __init__(self, first_row, agent_column, prechat_column, body_column, names_file=None,
                 drugs_file=None, agents_file= None, organisations_file=None, places_file=None, words_and_names_file=None):
        self.first_row = first_row
        self.agent_column = agent_column
        self.prechat_column = prechat_column
        self.body_column = body_column

        self.replace_lists = {}
        for type, file_name in {"name": names_file,
                                "organisation": organisations_file,
                                "place": places_file}.items():
            if file_name is not None:
                with open(file_name, encoding='UTF-8') as file:
                    lines = file.readlines()
                    lines = [line.strip() for line in lines]
                    self.replace_lists[type] = lines

        if drugs_file is not None:
            with open(drugs_file, encoding='UTF-8') as file:
                lines = file.readlines()
                lines = [line.strip() for line in lines]
                self.drugs_list = lines

        self.words_and_names = []
        if words_and_names_file is not None:
            with open(words_and_names_file, encoding='UTF-8') as words_and_names:
                self.words_and_names = [line.strip().lower() for line in words_and_names.readlines()]

        self.agent_dict = dict()
        self.agent_index = 0

        if agents_file is not None:
            with open(agents_file, encoding='UTF-8') as file:
                lines = file.readlines()
                lines = [line.strip() for line in lines]
                for line in lines:
                    self.add_agent_id(line)

    def process_file(self, file_name):
        """
        Processes one file.

        :param file_name:
        :return:
        """

        file_basename = os.path.basename(file_name)

        # try:
        # Load the workbook
        workbook = load_workbook(filename=file_name)
        # Take the first worksheet
        worksheet = workbook[workbook.get_sheet_names()[0]]

        for row_index in range(self.first_row, worksheet.max_row+1):
            try:
                # Agents
                agent = worksheet[self.agent_column+str(row_index)].value
                agent_name = ''
                agent_id = 'agent'
                if isinstance(agent, str):
                    agent_name = agent.capitalize()
                    agent_id = self.get_agent_id(agent_name)
                    worksheet[self.agent_column + str(row_index)] = "[" + agent_id + "]"

                # Prechat
                prechat = worksheet[self.prechat_column+str(row_index)].value
                user = ''
                if isinstance(prechat, str):

                    result = re.search(r'Naam:? = ?([^\n]*)\n', prechat)
                    if result:
                        user = result.group(1)

                    prechat = re.sub(r'(Naam:? = ?)[^\n]*(\n)', r'\1[name]\2', prechat)
                    prechat = re.sub(r'(Woonplaats:? = ?)[^\n]*(\n)', r'\1[place]\2', prechat)
                    prechat = re.sub(r'\d{4}[ \t]*[A-Za-z]{2}\b', '[zipcode]', prechat)
                    prechat = self.replace_using_list_items(prechat)
                    prechat = self.replace_email_address(prechat)
                    prechat = self.replace_url(prechat)
                    prechat = self.replace_acronyms(prechat)
                    worksheet[self.prechat_column + str(row_index)] = prechat

                if user is not '' and user == agent_name:
                    print("Warning: agent name and user name are equal (%s)." % agent_name)

                # Body
                body = worksheet[self.body_column+str(row_index)].value
                if isinstance(body, str):
                    # Get chat names from body
                    chat_names = set()
                    for match in re.compile(r'\] (\w+):').finditer(body):
                        chat_names.add(match.group(1))

                    # We expect 2 people in the chat
                    if len(chat_names) == 2 or (agent_name and user and agent_name != user):
                        if agent_name and agent_name in chat_names:
                            body = re.sub(r'(\] )(?!' + agent_name + r')\w+(:)', r'\1[name]\2', body)
                            body = re.sub(r'(\] )' + agent_name + r'(:)', r'\1[' + agent_id + r']\2', body)
                        elif user:
                            body = re.sub(r'(\] )' + user + r'(:)', r'\1[name]\2', body)
                            try:
                                chat_names_without_user = (chat_names - {user})
                                agent_name_from_body = chat_names_without_user.pop() if len(chat_names_without_user) else agent_name
                                if agent_name_from_body in self.agent_dict:
                                    agent_id_from_body = self.agent_dict[agent_name_from_body]
                                    body = re.sub(r'(\] )' + agent_name_from_body + r'(:)', r'\1[' + agent_id_from_body + r']\2', body)
                                else:
                                    body = re.sub(r'(\] )\w+(:)', r'\1[name]\2', body)
                            except Exception as e:
                                import traceback
                                traceback.print_exc()

                    # Replace any leftover names
                    body = re.sub(r'(\] )\w+(:)', r'\1[name]\2', body)

                    body = re.sub(r'\d{4}[ \t]*[A-Za-z]{2}\b', '[zipcode]', body)

                    body = self.replace_using_list_items(body)
                    body = self.replace_using_drugs_list(body)
                    body = self.replace_email_address(body)
                    body = self.replace_url(body)
                    body = self.replace_acronyms(body)

                    worksheet[self.body_column + str(row_index)] = body

            except Exception:
                print("Unable to process row " + str(row_index) + ". Error message: " + str(sys.exc_info()))

        workbook.save(self.output_dir + os.sep + file_basename)
        self.save_agent_ids()

        # except Exception as exception:
        #     print(exception)

    def add_agent_id(self, name):
        if name not in self.agent_dict:
            self.agent_index += 1
            agent_id = 'agent' + str(self.agent_index)
            self.agent_dict[name] = agent_id

    def get_agent_id(self, name):
        if name in self.agent_dict:
            return self.agent_dict[name]
        else:
            self.agent_index += 1
            agent_id = 'agent' + str(self.agent_index)
            self.agent_dict[name] = agent_id
            return agent_id

    def save_agent_ids(self):
        with open(self.output_dir + os.sep + self._agent_ids_csv, 'w') as agent_ids_file:
            for k,v in self.agent_dict.items():
                agent_ids_file.write("\"%s\",\"%s\"\n" % (k,v))


    def replace_using_list_items(self, text):
        for type, list in self.replace_lists.items():
            for item in list:
                if len(item) < 5 or item.lower() in self.words_and_names:
                    # Strict: NOT ignoring case
                    text = re.sub(r'\b' + re.sub(r'\\ ', '\s+', re.escape(item)) + r'\b', "[%s]" % type, text)
                else:
                    # Not strict: ignoring case
                    text = re.sub(r'\b' + re.sub(r'\\ ', '\s+', re.escape(item)) + r'\b', "[%s]" % type, text,
                                  flags=re.I)
        return text

    def replace_using_drugs_list(self, text):
        """Replace by [drug: <drug name>]"""
        for item in self.drugs_list:
            if len(item) < 3:
                new_text = re.sub(r'(?<!\[drug: )\b' + re.sub(r'\\ ', '\s+', re.escape(item)) + r'\b',
                                  "[%s: %s]" % ("drug", item), text)
            else:
                new_text = re.sub(r'(?<!\[drug: )' + re.sub(r'\\ ', '\s+', re.escape(item)),
                                  "[%s: %s]" % ("drug", item), text)
            if new_text != text:
                print("Replaced drug name: %s" % item, file=sys.stderr)
            text = new_text
        return text

    def replace_email_address(self, text):
        return re.sub(r"[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})", "[email]", text)

    def replace_url(self, text):
        return re.sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', "[url]", text)

    def replace_acronyms(self, text):
        pattern = r'\b[A-Z]{2,}\b'
        regex = re.compile(pattern)
        for match in regex.finditer(text):
            if match.group(0) not in self.drugs_list:
                text = re.sub(r'\b' + match.group(0) + r'\b', "[acronym]", text)
        return text
        # return re.sub(r'\b[A-Z]{2,}\b', "[acronym]", text)

    def get_extensions(self):
        return self._extensions

if __name__ == "__main__":
    # -o Output directory; optional
    usage = "Usage: \n" + sys.argv[0] + \
            " -o <output directory>" + \
            " -n <names list>" + \
            " -r <organisation list>" + \
            " -p <places list>" + \
            " -d <list of drugs/medicines>" + \
            " -w <list of names that are also words>" + \
            " -a <list of agents>"

    # Set default values
    output_dir = None

    # Register command line arguments
    opt_list, file_list = getopt.getopt(sys.argv[1:], 'o:n:r:p:d:w:a:')
    for opt in opt_list:
        if opt[0] == '-o':
            output_dir = opt[1]
        if opt[0] == '-n':
            names_file = opt[1]
        if opt[0] == '-r':
            organisations_file = opt[1]
        if opt[0] == '-p':
            places_file = opt[1]
        if opt[0] == '-d':
            drugs_file = opt[1]
        if opt[0] == '-a':
            agents_file = opt[1]
        if opt[0] == '-w':
            words_and_names_file = opt[1]

    # Check for errors and report
    errors = []
    if file_list is None or len(file_list) == 0:
        errors.append("No files or directories given.")

    if len(errors) != 0:
        print("Errors:")
        print("\n".join(errors))
        print(usage)
        sys.exit(1)

    # Report registered options
    print("OPTIONS", file=sys.stderr)
    print("Files: " + ", ".join(file_list), file=sys.stderr)
    print("Names file: " + names_file, file=sys.stderr)
    print("Organisations file: " + organisations_file, file=sys.stderr)
    print("Places file: " + places_file, file=sys.stderr)
    print("Drugs file: " + drugs_file, file=sys.stderr)
    print("Agents file: " + agents_file, file=sys.stderr)
    print("Words and names file: " + words_and_names_file, file=sys.stderr)
    if output_dir is not None:
        print("Output directory: " + output_dir, file=sys.stderr)

    # Build and run
    file_collection_processor = FileCollectionProcessor(file_list, output_dir=output_dir, extensions_to_process=["xlsx"])
    trimbosAnonymizer = TrimbosAnonymizer(13, "D", "W", "Y", names_file=names_file,
                                          organisations_file=organisations_file, places_file=places_file,
                                          drugs_file=drugs_file, agents_file=agents_file, words_and_names_file=words_and_names_file)
    file_collection_processor.add_file_processor(trimbosAnonymizer)
    file_collection_processor.run()