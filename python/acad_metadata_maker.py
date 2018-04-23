#!python

"""
TODO
"""

from jinja2 import Environment, select_autoescape, FileSystemLoader
from openpyxl import load_workbook
import sys
import os
from lxml import etree

if len(sys.argv) != 5:
    print("""Usage:
    %s <metadata-excel> <template-file> <output-dir> <folia-file-dir>
    """ % sys.argv[0])
    exit()


FIRST_ROW_IN_METADATA_XSLX = 2
NUMBER_OF_ROWS_TO_PROCESS = 4
METADATA_COLUMNS = [
    'file_name',
    'title',
    'authors',
    'date',
    'genre',
    'section', # Dutch: katern
    'medium',
    'url',
]
OUTPUT_FILE_EXTENSION = 'cmdi.xml'
FOLIA_FILE_EXTENSION = 'folia.xml'

metadata_excel_file = sys.argv[1]
template_file_path = sys.argv[2]
output_dir = sys.argv[3]
folia_dir = sys.argv[4]


def count_words_in_folia(folia_file):
    try:
        root = etree.parse(folia_file).getroot()
        count = root.xpath("count(//f:w[@class=\"WORD\" or @class=\"PUNCTUATION\" or @class=\"Vern\"])",
                           namespaces={'f':'http://ilk.uvt.nl/folia'})
        return int(count)
    except:
        sys.stderr.write("Folia file %s could not be analysed.\n" % folia_file)
        return None



# --- Prepare Jinja2 ---

template_file_name = os.path.basename(template_file_path)
template_dir_name = os.path.dirname(template_file_path)

env = Environment(
    autoescape=select_autoescape(['html', 'xml']),
    loader=FileSystemLoader(template_dir_name),
    trim_blocks=False
)
template = env.get_template(template_file_name)


# --- Read Metadata Excel ---

# Load the workbook
workbook = load_workbook(filename=metadata_excel_file)
# Take the first worksheet
worksheet = workbook[workbook.get_sheet_names()[0]]

# max_row = NUMBER_OF_ROWS_TO_PROCESS + FIRST_ROW_IN_METADATA_XSLX
max_row = worksheet.max_row+1

for row_index in range(FIRST_ROW_IN_METADATA_XSLX, max_row):
    row_index_str = str(row_index)

    # Put the data of a row in a convenient dictionary
    row_data = [cell.value for cell in worksheet['A'+row_index_str:'H'+row_index_str][0]]
    row_dict = dict(list(zip(METADATA_COLUMNS, row_data)))

    file_name = row_dict['file_name']
    word_count = count_words_in_folia(os.path.sep.join((folia_dir, file_name + "." + FOLIA_FILE_EXTENSION)))
    if word_count is None:
        continue
        
    print('Medium: "{0}"'.format(row_dict['medium']))
    if row_dict['medium'] == 'Digitaal':
        text_type = 'WR-P-E-M_newspapers'
    elif row_dict['medium'] == 'Papieren krant':
        text_type = 'WR-P-P-G_newspapers'
    else:
        text_type = ''

    output_file = os.path.sep.join((output_dir, file_name + '.' + OUTPUT_FILE_EXTENSION))

    with open(output_file, 'w') as file:
        file.write(template.render(
            project_name='ACAD',
            text_type=text_type,
            published='Y',
            publisher='NRC Handelsblad',
            publication_date_yyyy_mm_dd=row_dict['date'].strftime("%Y-%m-%d"),
            genre=row_dict['genre'],
            author_name=row_dict['authors'],
            text_title=row_dict['title'],
            publication_section=row_dict['section'] or '',
            resource_proxy_id=row_dict['file_name'],
            number_of_word_tokens=word_count
        ))
        file.close()
