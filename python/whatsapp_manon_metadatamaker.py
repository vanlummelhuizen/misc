#!python

"""
TODO
"""

from jinja2 import Environment, select_autoescape, FileSystemLoader
from openpyxl import load_workbook
import sys
import os
from lxml import etree
from unidecode import unidecode

if len(sys.argv) != 6:
    print("""Usage:
    %s <metadata-excel> <template-file> <authors-template-file> <output-dir> <folia-file-dir>
    """ % sys.argv[0])
    exit()


FIRST_ROW_IN_METADATA_XSLX = 2
NUMBER_OF_ROWS_TO_PROCESS = 4
METADATA_COLUMNS = [
    'user_id',
    'name',
    'file_name',
    'nickname',
    'birthdate',
    'age',
    'residence',
    'highest_education',
    'supplier_name',
    'signed_date'
]
OUTPUT_FILE_EXTENSION = 'cmdi.xml'
FOLIA_FILE_EXTENSION = 'folia.xml'

# metadata_excel_file = sys.argv[1]
# template_file_path = sys.argv[2]
# authors_file_path = sys.argv[3]
# output_dir = sys.argv[4]
# folia_dir = sys.argv[5]

metadata_excel_file, template_file_path, authors_template_file_path, output_dir, folia_dir = sys.argv[1:]


def count_words_in_folia(folia_file):
    print(folia_file)
    try:
        xmlparser = etree.XMLParser(encoding='ISO-8859-1')
        root = etree.parse(folia_file, xmlparser).getroot()
        count = root.xpath("count(//f:w[@class=\"WORD\" or @class=\"PUNCTUATION\" or @class=\"Vern\"])",
                           namespaces={'f':'http://ilk.uvt.nl/folia'})
        return int(count)
    except:
        sys.stderr.write("Folia file %s could not be analysed.\n" % folia_file)
        sys.stderr.write(str(sys.exc_info()))
        return None



# --- Prepare Jinja2 ---

template_file_name = os.path.basename(template_file_path)
template_dir_name = os.path.dirname(template_file_path)
authors_template_file_name = os.path.basename(authors_template_file_path)

env = Environment(
    autoescape=select_autoescape(['html', 'xml']),
    loader=FileSystemLoader(template_dir_name),
    trim_blocks=False
)
template = env.get_template(template_file_name)
# authors_template = env.get_template(authors_template_file_name)


# --- Read Metadata Excel ---

# Load the workbook
workbook = load_workbook(filename=metadata_excel_file)
# Take the first worksheet
worksheet = workbook[workbook.get_sheet_names()[0]]

# max_row = NUMBER_OF_ROWS_TO_PROCESS + FIRST_ROW_IN_METADATA_XSLX
max_row = worksheet.max_row+1

# First load all data from the metadata xslx
data_dict = {}
for row_index in range(FIRST_ROW_IN_METADATA_XSLX, max_row):
    row_index_str = str(row_index)

    # Put the data of a row in a convenient dictionary
    row_data = [cell.value for cell in worksheet['A'+row_index_str:'J'+row_index_str][0]]
    row_dict = dict(list(zip(METADATA_COLUMNS, row_data)))
    row_dict['file_name'] = row_dict['file_name'].replace(" ", "-")
    row_dict['file_name'] = row_dict['file_name'].replace("(", "")
    row_dict['file_name'] = row_dict['file_name'].replace(")", "")
    row_dict['file_name'] = row_dict['file_name'].replace("[", "")
    row_dict['file_name'] = row_dict['file_name'].replace("]", "")
    row_dict['file_name'] = unidecode(row_dict['file_name'])

    file_name = row_dict['file_name']
    if file_name != "-":
        if file_name in data_dict:
            data_dict[file_name].append(row_dict)
        else:
            data_dict[file_name] = [row_dict]

for file_name in data_dict:
    file_name_without_extension = os.path.splitext(file_name)[0]
    word_count = count_words_in_folia(os.path.sep.join((folia_dir, file_name_without_extension + "." + FOLIA_FILE_EXTENSION)))
    if not word_count:
        continue

    output_file = os.path.sep.join((output_dir, file_name + '.' + OUTPUT_FILE_EXTENSION))

    authors = []
    for user_dict in data_dict[file_name]:
        authors.append({
            'pseudonym':user_dict['user_id'],
            'age':user_dict['age'],
            'residence_country':"The Netherlands",
            'residence_town':user_dict['residence']
        })


    with open(output_file, 'w') as file:
        file.write(template.render(
            project_name='ACAD',
            text_type='WR-U-E-A_chats',
            published='N',
            publisher='',
            # publication_date_yyyy_mm_dd=row_dict['date'].strftime("%Y-%m-%d"),
            # genre=row_dict['genre'],
            # author_name=row_dict['authors'],
            authors=authors,
            resource_proxy_id=file_name,
            number_of_word_tokens=word_count or ''
        ))
        file.close()
